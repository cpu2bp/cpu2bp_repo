import json
import openpyxl
import re

excel_path = r"C:\CPU2BP\CSR\MMC_5K45_CSR_CPD_Test_Catalogue_measurement_campaign.xlsx"
json_path = r"C:\CPU2BP\CSR\MMC\newjson.json"

# Mapping rules
def occupant_to_code(val):
    if not val or str(val).strip().lower() in ["", "none"]:
        return "E"
    val = str(val).strip()
    if val.lower() in ["1yo", "3yo", "6yo", "new born"]:
        return "C"
    if val.lower() in ["static toy", "moving toy", "bag"]:
        return "O"
    if val.lower() == "liquid":
        return "W"
    if val.lower() == "phone":
        return "P"
    if val.upper() in ["AM%50", "AF%05"]:
        return "A"
    return "E"

def occupant_to_code_footwell(val):
    if not val or str(val).strip().lower() in ["", "none"]:
        return "E"
    val = str(val).strip()
    if val.lower() in ["1yo", "3yo", "6yo", "new born"]:
        return "CF"
    if val.lower() in ["static toy", "moving toy", "bag"]:
        return "OF"
    if val.lower() == "liquid":
        return "WF"
    if val.lower() == "phone":
        return "PF"
    if val.upper() in ["AM%50", "AF%05"]:
        return "AF"
    return "E"

seat_map = [
    ("Driver Seat Occupant", "1L"),
    ("Passenger Seat Occupant", "1R"),
    ("2nd Rear-L Seat Occupant", "2L"),
    ("2nd Rear-C Seat Occupant", "2M"),
    ("2nd Rear-R Seat Occupant", "2R"),
    ("3rd Rear-L Seat Occupant", "3L"),
    ("3rd Rear-R Seat Occupant", "3R"),
]

footwell_map = [
    ("Driver  Footwell Occupant", "1L"),
    ("Passenger  Footwell Occupant", "1R"),
    ("2nd Rear-L Footwell Occupant", "2L"),
    ("2nd Rear-C Footwell Occupant", "2M"),
    ("2nd-Rear-R Footwell Occupant", "2R"),
    ("3rd Rear-L Seat Footwell Occupant", "3L"),
    ("3rd-Rear-R  Footwell Occupant", "3R"),
]

# Load Excel
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb.active

# Find header row and build column index
header_row = None
header_idx = {}
for row in ws.iter_rows(min_row=1, max_row=10):
    for cell in row:
        if cell.value and str(cell.value).strip() == "Scenario No.":
            header_row = cell.row
            break
    if header_row:
        break

if not header_row:
    raise Exception("Header row not found in Excel.")

for cell in ws[header_row]:
    if cell.value:
        header_idx[str(cell.value).strip()] = cell.column

print("Found headers:", list(header_idx.keys()))

# Build a mapping from scenario number to row
scenario_map = {}
for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
    scenario_cell = row[header_idx["Scenario No."]-1]
    if scenario_cell.value is not None:
        scenario_map[str(scenario_cell.value).strip()] = row

# Load JSON
with open(json_path, "r", encoding="utf-8") as f:
    data = json.load(f)

for entry in data:
    name = entry.get("name", "")
    # Match TC, tc, TC_, tc_ in name (case-insensitive)
    #if re.search(r'(TC|tc|TC_|tc_)', name):
    for folder in entry.get("folders", []):
        files = folder.get("files", [])
        tc_number = None
        for fname in files:
            m = re.search(r"TC[_]?(\d+)", fname)
            if m:
                tc_number = m.group(1)
                break
        if not tc_number:
            continue
        # Find Excel row for this TC number
        row = scenario_map.get(tc_number)
        if not row:
            print(f"TC Number {tc_number} not found in Excel.")
            continue
        sod_gt = {}
        for seat_col, sod_key in seat_map:
            footwell_col = None
            # Find corresponding footwell column
            for f_col, f_key in footwell_map:
                if f_key == sod_key:
                    footwell_col = f_col
                    break

            footwell_val = None
            if footwell_col:
                footwell_col_idx = header_idx.get(footwell_col)
                if footwell_col_idx:
                    footwell_val = row[footwell_col_idx-1].value

            if footwell_val and str(footwell_val).strip().lower() not in ["", "none"]:
                sod_gt[sod_key] = occupant_to_code_footwell(footwell_val)
            else:
                seat_col_idx = header_idx.get(seat_col)
                if seat_col_idx:
                    seat_val = row[seat_col_idx-1].value
                    sod_gt[sod_key] = occupant_to_code(seat_val)
                else:
                    sod_gt[sod_key] = "E"

        print(f"Updating TC {tc_number} with SOD_GT: {sod_gt}")

        gt = folder.get("gt", {})
        gt["SOD_GT"] = sod_gt
        folder["gt"] = gt

# Debugging for TC 138
tc_138_row = scenario_map.get("138")
if tc_138_row:
    driver_footwell_col_idx = header_idx.get("Driver  Footwell Occupant")
    driver_seat_col_idx = header_idx.get("Driver Seat Occupant")
    
    if driver_footwell_col_idx:
        driver_footwell_val = tc_138_row[driver_footwell_col_idx-1].value
        print(f"Debug TC 138 - Driver Footwell Occupant: {driver_footwell_val}")
    else:
        print("Debug TC 138 - 'Driver  Footwell Occupant' column not found.")
        
    if driver_seat_col_idx:
        driver_seat_val = tc_138_row[driver_seat_col_idx-1].value
        print(f"Debug TC 138 - Driver Seat Occupant: {driver_seat_val}")
    else:
        print("Debug TC 138 - 'Driver Seat Occupant' column not found.")


with open(json_path, "w", encoding="utf-8") as f:
    json.dump(data, f, indent=4, ensure_ascii=False)

print("JSON updated from Excel.")