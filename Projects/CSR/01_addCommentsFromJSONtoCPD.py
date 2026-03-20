import json
import openpyxl
import os

def update_excel_with_comments(json_path, excel_path):
    """
    Reads a JSON file for measurement data and updates an Excel file with comments,
    only if the target cell is empty. It prioritizes non-null comments if duplicates are found.

    Args:
        json_path (str): The file path for the JSON input file.
        excel_path (str): The file path for the Excel file to be updated.
    """
    # --- Load JSON file ---
    try:
        with open(json_path, 'r') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"Error: JSON file not found at {json_path}")
        return
    except json.JSONDecodeError:
        print(f"Error: Could not decode JSON from {json_path}")
        return

    # --- Load Excel file ---
    try:
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
    except FileNotFoundError:
        print(f"Error: Excel file not found at {excel_path}")
        return

    # --- Find column indices ---
    header = [cell.value for cell in sheet[1]]
    try:
        measurement_col_idx = header.index("Measurement name") + 1
        comment_col_idx = header.index("JsonComment") + 1
    except ValueError as e:
        print(f"Error: A required column was not found in the Excel file - {e}")
        print(f"Available columns are: {header}")
        return

    # --- Create mapping from JSON, prioritizing non-null comments ---
    measurement_comments = {}
    for entry in data:
        if "folders" in entry:
            for folder in entry["folders"]:
                if "files" in folder:
                    for filename, comment in folder["files"].items():
                        # If the filename is already stored, only overwrite it if the new comment is not null.
                        # If the filename is not stored, add it regardless of the comment value.
                        if filename not in measurement_comments or comment is not None:
                            measurement_comments[filename] = comment

    # --- Iterate through Excel and update comments ---
    updated_count = 0
    skipped_count = 0
    for row in range(2, sheet.max_row + 1):
        measurement_name_cell = sheet.cell(row=row, column=measurement_col_idx)
        measurement_name = measurement_name_cell.value

        if measurement_name and measurement_name in measurement_comments:
            comment_to_add = measurement_comments[measurement_name]
            
            # Proceed only if the comment from JSON is not null
            if comment_to_add is not None:
                comment_cell = sheet.cell(row=row, column=comment_col_idx)
                
                # Only update if the Excel cell is empty
                if not comment_cell.value:
                    comment_cell.value = comment_to_add
                    updated_count += 1
                else:
                    skipped_count += 1

    if updated_count > 0:
        print(f"Found and updated {updated_count} empty cells.")
    if skipped_count > 0:
        print(f"Skipped {skipped_count} cells that already had comments.")
    if updated_count == 0 and skipped_count == 0:
        print("No matching measurement names found to update.")

    # --- Save the changes ---
    if updated_count > 0:
        try:
            workbook.save(excel_path)
            print(f"Successfully saved updates to {excel_path}")
        except Exception as e:
            print(f"Error saving Excel file: {e}")
    else:
        print("No changes were made, so the file was not saved.")

if __name__ == "__main__":
    json_file = 'measurement_catalog.json'
    excel_file = 'CPD_measurement_catalogue.xlsx'
    
    if not os.path.exists(json_file):
        print(f"Error: '{json_file}' not found in the current directory.")
    elif not os.path.exists(excel_file):
        print(f"Error: '{excel_file}' not found in the current directory.")
    else:
        update_excel_with_comments(json_file, excel_file)
